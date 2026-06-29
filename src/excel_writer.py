from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .config import Config


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, header=True):
    rows = dataframe_to_rows(df, index=False, header=header)
    r = start_row
    for row in rows:
        c = start_col
        for v in row:
            ws.cell(row=r, column=c, value=v)
            c += 1
        r += 1


def write_excel_report(df_flagged: pd.DataFrame, aggregates: Dict[str, pd.DataFrame], cfg: Config) -> Path:
    out_path: Path = cfg.data_processed_excel
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # default sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # ---- Summary sheet ----
    ws_summary["A1"] = "Revenue Leakage QMR"
    ws_summary["A1"].font = Font(size=14, bold=True)

    kpi = aggregates["kpi"]
    ws_summary.append([])
    # write KPI table
    _write_df(ws_summary, kpi, start_row=3, start_col=1, header=True)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # style header rows for KPI
    for cell in ws_summary[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # by month
    ws_month = wb.create_sheet("Leakage_By_Month")
    ws_month["A1"] = "Leakage mismatches by month and currency"
    ws_month["A1"].font = Font(bold=True)

    by_month = aggregates["by_month"]
    _write_df(ws_month, by_month, start_row=3, start_col=1, header=True)
    for cell in ws_month[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # by quarter
    ws_q = wb.create_sheet("Leakage_By_Quarter")
    ws_q["A1"] = "Leakage mismatches by quarter and currency"
    ws_q["A1"].font = Font(bold=True)

    by_quarter = aggregates["by_quarter"]
    _write_df(ws_q, by_quarter, start_row=3, start_col=1, header=True)
    for cell in ws_q[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # mismatch rows
    ws_rows = wb.create_sheet("Leakage_Rows")
    ws_rows["A1"] = "Mismatching rows (flagged)"
    ws_rows["A1"].font = Font(bold=True)

    mismatch_rows = df_flagged[df_flagged["is_mismatch"]].copy()
    mismatch_rows = mismatch_rows.sort_values("abs_delta_amt", ascending=False)

    # Keep relevant columns
    cols = [
        cfg.col_transaction_date,
        cfg.col_currency,
        cfg.col_customer_id,
        cfg.col_invoice_id,
        cfg.col_transaction_amt,
        cfg.col_reported_amt,
        "delta_amt",
        "abs_delta_amt",
        "leakage_direction",
        "year_month",
        "quarter",
    ]
    mismatch_rows = mismatch_rows[cols]

    _write_df(ws_rows, mismatch_rows, start_row=3, start_col=1, header=True)
    for cell in ws_rows[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # top invoices
    ws_top = wb.create_sheet("Top_Invoices")
    ws_top["A1"] = "Top mismatches by absolute delta"
    ws_top["A1"].font = Font(bold=True)

    top = aggregates["top_invoices"]
    _write_df(ws_top, top, start_row=3, start_col=1, header=True)
    for cell in ws_top[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Data quality sheet (lightweight)
    ws_dq = wb.create_sheet("Data_Quality")
    ws_dq["A1"] = "Data validation summary"
    ws_dq["A1"].font = Font(bold=True)

    validation = df_flagged.attrs.get("validation", {}) or {}

    # Ensure we only pass Excel-serializable values (avoid nested dicts)
    items = []
    for k, v in validation.items():
        if isinstance(v, dict):
            items.append((k, str(v)))
        else:
            items.append((k, v))

    dq_df = pd.DataFrame(items, columns=["metric", "value"])
    if dq_df.empty:
        dq_df = pd.DataFrame([{"metric": "note", "value": "No validation metadata available."}])


    _write_df(ws_dq, dq_df, start_row=3, start_col=1, header=True)
    for cell in ws_dq[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # basic freeze panes + number formats
    for ws in [ws_summary, ws_month, ws_q, ws_rows, ws_top, ws_dq]:
        ws.freeze_panes = "A4"

    # Save
    wb.save(out_path)
    return out_path

